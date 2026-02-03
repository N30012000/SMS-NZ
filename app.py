import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image
import pdf2image
import cv2
import pytesseract
from paddleocr import PaddleOCR
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import tempfile
import os
import json
from typing import Dict, List, Optional, Tuple
import warnings
warnings.filterwarnings('ignore')

# Initialize PaddleOCR (free, no API limits)
ocr = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)

# Initialize session state
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = []
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = pd.DataFrame()

# Standardized values for dropdowns
STANDARDIZED_VALUES = {
    'locations': [
        'Ramp', 'Galley', 'Cabin', 'Flight Deck', 
        'Remote Parking Bay', 'Terminal', 'Baggage Area',
        'Cargo Bay', 'Maintenance Hangar', 'Fueling Area',
        'Catering', 'Other'
    ],
    'hazard_types': [
        'Operational', 'Ground Handling', 'Cabin Safety',
        'Maintenance', 'Human Factors', 'Security',
        'Environmental', 'Technical', 'Procedural',
        'Equipment Failure', 'Other'
    ],
    'departments': [
        'Airport Services', 'Flight Operations', 'Maintenance',
        'Cabin Crew', 'Ground Handling', 'Security',
        'Quality & Safety', 'Catering', 'Engineering',
        'Other'
    ],
    'severity_levels': [
        'Catastrophic', 'Major', 'Moderate', 'Minor', 'Insignificant'
    ],
    'probability_levels': [
        'Frequent', 'Occasional', 'Remote', 'Improbable', 'Rare'
    ],
    'risk_levels': ['High', 'Medium', 'Low'],
    'yes_no': ['Yes', 'No']
}

def preprocess_image(image):
    """Preprocess image for better OCR"""
    # Convert to grayscale
    if len(image.shape) == 3:
        gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    else:
        gray = np.array(image)
    
    # Apply thresholding
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Remove noise
    kernel = np.ones((1,1), np.uint8)
    processed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
    
    return Image.fromarray(processed)

def extract_text_with_paddleocr(image):
    """Extract text using PaddleOCR"""
    try:
        # Convert PIL to numpy array
        img_array = np.array(image)
        
        # Run OCR
        result = ocr.ocr(img_array, cls=True)
        
        # Extract text
        all_text = []
        if result[0] is not None:
            for line in result[0]:
                text = line[1][0]
                confidence = line[1][1]
                if confidence > 0.5:  # Filter by confidence
                    all_text.append(text)
        
        return " ".join(all_text)
    except Exception as e:
        st.warning(f"OCR Error: {str(e)}")
        return ""

def parse_date(date_str):
    """Parse and normalize dates to DD-MM-YYYY"""
    date_formats = [
        '%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y',
        '%Y-%m-%d', '%Y/%m/%d',
        '%d-%b-%Y', '%d %b %Y',
        '%d-%B-%Y', '%d %B %Y'
    ]
    
    for fmt in date_formats:
        try:
            date_obj = datetime.strptime(date_str.strip(), fmt)
            return date_obj.strftime('%d-%m-%Y')
        except:
            continue
    
    return "Unclear"

def extract_form_fields(text):
    """Extract all required fields from form text"""
    data = {
        'basic_details': {},
        'hazard_details': {},
        'initial_risk': {},
        'cap': {},
        'residual_risk': {},
        'administrative': {}
    }
    
    lines = text.split('\n')
    
    # Extract Report Number
    report_no_patterns = [
        r'Report\s*(?:no\.?|number|#)[:\s]*([A-Za-z0-9\-/]+)',
        r'Tracking\s*#\s*[:]?\s*([A-Za-z0-9\-/]+)',
        r'HR-Con-[-\s]*([0-9/]+)',
        r'1H1-ESR-3C/(\d{4})'
    ]
    
    for pattern in report_no_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['basic_details']['report_number'] = match.group(1).strip()
            break
    else:
        data['basic_details']['report_number'] = "Unclear"
    
    # Extract Date of Report
    date_patterns = [
        r'Date\s*of\s*Report[:\s]*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})',
        r'(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})\s*Date of Report'
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['basic_details']['date_of_report'] = parse_date(match.group(1))
            break
    else:
        data['basic_details']['date_of_report'] = "Unclear"
    
    # Extract Reporter Name
    name_patterns = [
        r'Reporter\s*Name[:\s]*([A-Za-z\s]+)(?:\s*Contact|\s*Department|$)',
        r'Name\s*[&]?\s*Contact[:\s]*([A-Za-z\s]+)'
    ]
    
    for pattern in name_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['basic_details']['reporter_name'] = match.group(1).strip()
            break
    else:
        data['basic_details']['reporter_name'] = "Unclear"
    
    # Extract Department
    dept_patterns = [
        r'Department[:\s]*([A-Za-z\s&]+)',
        r'Airport\s*Services',
        r'Flight\s*Operations',
        r'Maintenance'
    ]
    
    for pattern in dept_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            dept = match.group(1).strip() if match.groups() else pattern
            data['basic_details']['department'] = dept
            break
    else:
        data['basic_details']['department'] = "Unclear"
    
    # Extract Location
    location_patterns = [
        r'Location\s*of\s*Hazard[:\s]*([A-Za-z\s]+)',
        r'Ramp\s*\[?([A-Za-z\s/]+)\]?',
        r'Remote\s*Parking\s*Bays?'
    ]
    
    for pattern in location_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            loc = match.group(1).strip() if match.groups() else pattern
            data['basic_details']['location'] = loc
            break
    else:
        data['basic_details']['location'] = "Unclear"
    
    # Extract Date/Time Hazard Identified
    hazard_date_patterns = [
        r'Date/Time\s*Hazard\s*Identified[:\s]*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})',
        r'Hazard\s*Identified[:\s]*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})'
    ]
    
    for pattern in hazard_date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['hazard_details']['hazard_datetime'] = parse_date(match.group(1))
            break
    else:
        data['hazard_details']['hazard_datetime'] = "Unclear"
    
    # Extract Hazard Description (simple extraction)
    # Look for sections after "Hazard Identification" or before "Initial Risk"
    hazard_section_patterns = [
        r'Hazard\s*Identification[\s:]+(.+?)(?:Initial Risk|Corrective Action|Severity|$)',
        r'Description[:\s]+(.+?)(?:Root Cause|Remarks|$)'
    ]
    
    for pattern in hazard_section_patterns:
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            desc = match.group(1).strip()
            if len(desc) > 10:  # Ensure meaningful content
                data['hazard_details']['hazard_description'] = desc[:500]  # Limit length
                break
    
    if 'hazard_description' not in data['hazard_details']:
        data['hazard_details']['hazard_description'] = "Unclear"
    
    # Extract Remarks
    remarks_patterns = [
        r'Remarks[:\s]+(.+?)(?:Sign-off|$)',
        r'Remarks\s*\(if any\)[:\s]+(.+?)(?:Sign-off|$)'
    ]
    
    for pattern in remarks_patterns:
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            data['hazard_details']['remarks'] = match.group(1).strip()[:300]
            break
    else:
        data['hazard_details']['remarks'] = ""
    
    # Extract Risk Levels (simplified detection)
    risk_keywords = {
        'severity': ['Catastrophic', 'Major', 'Moderate', 'Minor', 'Insignificant'],
        'probability': ['Frequent', 'Occasional', 'Remote', 'Improbable', 'Rare'],
        'risk_level': ['High', 'Medium', 'Low']
    }
    
    for risk_type, keywords in risk_keywords.items():
        for keyword in keywords:
            if keyword.lower() in text.lower():
                if risk_type == 'severity':
                    data['initial_risk'][risk_type] = keyword
                elif risk_type == 'probability':
                    data['initial_risk'][risk_type] = keyword
                elif risk_type == 'risk_level':
                    # Check context for initial vs residual
                    if 'Initial Risk' in text or 'Initial' in text:
                        data['initial_risk'][risk_type] = keyword
                    elif 'Residual' in text or 'After' in text:
                        data['residual_risk'][risk_type] = keyword
    
    # Set defaults if not found
    if 'severity' not in data['initial_risk']:
        data['initial_risk']['severity'] = "Unclear"
    if 'probability' not in data['initial_risk']:
        data['initial_risk']['probability'] = "Unclear"
    if 'risk_level' not in data['initial_risk']:
        data['initial_risk']['risk_level'] = "Unclear"
    
    # Extract CAP information
    cap_patterns = [
        r'Action\s*Plan[:\s]+(.+?)(?:Target Date|$)',
        r'Corrective\s*Action[:\s]+(.+?)(?:Target Date|Responsible Person|$)',
        r'CAP[:\s]+(.+?)(?:Target Date|$)'
    ]
    
    for pattern in cap_patterns:
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            data['cap']['action_description'] = match.group(1).strip()[:500]
            data['cap']['cap_required'] = "Yes"
            break
    else:
        data['cap']['action_description'] = ""
        data['cap']['cap_required'] = "No"
    
    # Extract Target Date
    target_date_patterns = [
        r'Target\s*Date[:\s]*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})',
        r'Completion\s*Date[:\s]*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})'
    ]
    
    for pattern in target_date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['cap']['target_date'] = parse_date(match.group(1))
            break
    else:
        data['cap']['target_date'] = ""
    
    # Extract Responsible Person
    resp_person_patterns = [
        r'Responsible\s*Person[:\s]*([A-Za-z\s.]+)',
        r'Person\s*Responsible[:\s]*([A-Za-z\s.]+)',
        r'Ms\.\s*([A-Za-z\s.]+)',
        r'Mr\.\s*([A-Za-z\s.]+)'
    ]
    
    for pattern in resp_person_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['cap']['responsible_person'] = match.group(1).strip()
            break
    else:
        data['cap']['responsible_person'] = ""
    
    # Extract Responsible Department
    resp_dept_patterns = [
        r'Responsible\s*Department[:\s]*([A-Za-z\s&]+)',
        r'Department[:\s]*([A-Za-z\s&]+)\s*(?:Target|$)'
    ]
    
    for pattern in resp_dept_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['cap']['responsible_department'] = match.group(1).strip()
            break
    else:
        data['cap']['responsible_department'] = ""
    
    # Check for CAP Attachment
    data['cap']['cap_attachment'] = "Yes" if "CAP is attached" in text or "attached" in text.lower() else "No"
    
    # Check for Wet Lease
    data['administrative']['wet_lease'] = "Yes" if "wet lease" in text.lower() or "wet-leased" in text.lower() else "No"
    
    # Extract Operator Name
    operator_patterns = [
        r'Operator[:\s]*([A-Za-z\s]+)',
        r'Aircraft\s*Operator[:\s]*([A-Za-z\s]+)'
    ]
    
    for pattern in operator_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['administrative']['operator'] = match.group(1).strip()
            break
    else:
        data['administrative']['operator'] = ""
    
    # Check if report attached
    data['administrative']['report_attached'] = "Yes" if "attached" in text.lower() else "No"
    
    # Determine Risk Status
    if data['residual_risk'].get('risk_level') == "High":
        data['residual_risk']['risk_status'] = "Intolerable"
    elif data['residual_risk'].get('risk_level') == "Medium":
        data['residual_risk']['risk_status'] = "Tolerable"
    elif data['residual_risk'].get('risk_level') == "Low":
        data['residual_risk']['risk_status'] = "Acceptable"
    else:
        data['residual_risk']['risk_status'] = "Unclear"
    
    return data

def create_excel_template():
    """Create audit-ready Excel template with all sheets"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Sheet 1: Raw SMS Data
    ws1 = wb.create_sheet(title="Raw SMS Data")
    
    headers1 = [
        'Report No', 'Date of Report', 'Reporter Name', 'Department',
        'Location', 'Hazard Type', 'Description', 'Hazard Date/Time',
        'Severity (Initial)', 'Probability (Initial)', 'Initial Risk Level',
        'CAP Required', 'CAP Description', 'Responsible Person',
        'Responsible Dept', 'Target Date', 'CAP Attachment',
        'Residual Severity', 'Residual Probability', 'Residual Risk Level',
        'Risk Status', 'Wet Lease', 'Aircraft Type', 'Operator',
        'Remarks', 'Status', 'Days Overdue'
    ]
    
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = [15, 15, 20, 20, 20, 20, 40, 20, 
                    15, 15, 15, 12, 40, 20, 20, 15, 
                    15, 15, 15, 15, 15, 12, 15, 20, 
                    30, 12, 12]
    
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # Sheet 2: Standardized Lists
    ws2 = wb.create_sheet(title="Standardized Lists")
    
    standardized_data = {
        'A1': 'Locations',
        'A2': 'Hazard Types', 
        'A3': 'Departments',
        'A4': 'Severity Levels',
        'A5': 'Probability Levels',
        'A6': 'Risk Levels',
        'A7': 'Yes/No',
        'A8': 'Status Options'
    }
    
    for cell_ref, title in standardized_data.items():
        ws2[cell_ref] = title
        ws2[cell_ref].font = Font(bold=True)
    
    # Populate lists
    lists = {
        'B': STANDARDIZED_VALUES['locations'],
        'C': STANDARDIZED_VALUES['hazard_types'],
        'D': STANDARDIZED_VALUES['departments'],
        'E': STANDARDIZED_VALUES['severity_levels'],
        'F': STANDARDIZED_VALUES['probability_levels'],
        'G': STANDARDIZED_VALUES['risk_levels'],
        'H': STANDARDIZED_VALUES['yes_no'],
        'I': ['Open', 'Closed', 'In Progress']
    }
    
    for col_letter, values in lists.items():
        for i, value in enumerate(values, 2):
            ws2[f'{col_letter}{i}'] = value
    
    # Sheet 3: CAP Tracker
    ws3 = wb.create_sheet(title="CAP Tracker")
    
    headers3 = [
        'Report No', 'Hazard Description', 'Responsible Person',
        'Responsible Dept', 'Target Date', 'Status', 'Days Overdue',
        'Traffic Light'
    ]
    
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Sheet 4: Monthly Dashboard (Placeholder)
    ws4 = wb.create_sheet(title="Monthly Dashboard")
    ws4['A1'] = "Monthly Safety Performance Dashboard"
    ws4['A1'].font = Font(size=16, bold=True, color="366092")
    
    # Sheet 5: Audit Evidence Log
    ws5 = wb.create_sheet(title="Audit Evidence Log")
    
    headers5 = [
        'Report No', 'CAP Attachment Available', 'Evidence File Name',
        'Date Verified', 'Verified By', 'Auditor Remarks', 'Status'
    ]
    
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    return wb

def add_data_to_excel(wb, data):
    """Add extracted data to Excel template"""
    ws1 = wb['Raw SMS Data']
    
    # Find next empty row
    next_row = ws1.max_row + 1
    
    # Prepare row data
    row_data = [
        data['basic_details'].get('report_number', ''),
        data['basic_details'].get('date_of_report', ''),
        data['basic_details'].get('reporter_name', ''),
        data['basic_details'].get('department', ''),
        data['basic_details'].get('location', ''),
        "Operational",  # Default hazard type
        data['hazard_details'].get('hazard_description', ''),
        data['hazard_details'].get('hazard_datetime', ''),
        data['initial_risk'].get('severity', ''),
        data['initial_risk'].get('probability', ''),
        data['initial_risk'].get('risk_level', ''),
        data['cap'].get('cap_required', ''),
        data['cap'].get('action_description', ''),
        data['cap'].get('responsible_person', ''),
        data['cap'].get('responsible_department', ''),
        data['cap'].get('target_date', ''),
        data['cap'].get('cap_attachment', ''),
        data['residual_risk'].get('severity', ''),
        data['residual_risk'].get('probability', ''),
        data['residual_risk'].get('risk_level', ''),
        data['residual_risk'].get('risk_status', ''),
        data['administrative'].get('wet_lease', ''),
        "",  # Aircraft Type
        data['administrative'].get('operator', ''),
        data['hazard_details'].get('remarks', ''),
        "Open",  # Default status
        0  # Days Overdue
    ]
    
    # Write data
    for col, value in enumerate(row_data, 1):
        ws1.cell(row=next_row, column=col, value=value)
    
    # Update CAP Tracker
    ws3 = wb['CAP Tracker']
    next_row_tracker = ws3.max_row + 1
    
    # Calculate days overdue
    days_overdue = 0
    status = "Open"
    traffic_light = "🟢"
    
    target_date = data['cap'].get('target_date', '')
    if target_date:
        try:
            target_dt = datetime.strptime(target_date, '%d-%m-%Y')
            days_overdue = (datetime.now() - target_dt).days
            if days_overdue > 0:
                status = "Overdue"
                traffic_light = "🔴"
            elif days_overdue >= -7:
                status = "Due Soon"
                traffic_light = "🟡"
        except:
            pass
    
    tracker_data = [
        data['basic_details'].get('report_number', ''),
        data['hazard_details'].get('hazard_description', '')[:100],
        data['cap'].get('responsible_person', ''),
        data['cap'].get('responsible_department', ''),
        target_date,
        status,
        max(0, days_overdue),
        traffic_light
    ]
    
    for col, value in enumerate(tracker_data, 1):
        ws3.cell(row=next_row_tracker, column=col, value=value)
    
    # Update Audit Evidence Log
    ws5 = wb['Audit Evidence Log']
    next_row_audit = ws5.max_row + 1
    
    audit_data = [
        data['basic_details'].get('report_number', ''),
        data['cap'].get('cap_attachment', ''),
        f"Report_{data['basic_details'].get('report_number', '')}.pdf",
        datetime.now().strftime('%d-%m-%Y'),
        "System",
        "Automatically logged by SMS Scanner",
        "Pending Review"
    ]
    
    for col, value in enumerate(audit_data, 1):
        ws5.cell(row=next_row_audit, column=col, value=value)
    
    return wb

def create_dashboard(df, month, year):
    """Create interactive dashboard from data"""
    
    # Filter data for selected month/year
    df['Date of Report'] = pd.to_datetime(df['Date of Report'], errors='coerce', dayfirst=True)
    filtered_df = df[
        (df['Date of Report'].dt.month == month) &
        (df['Date of Report'].dt.year == year)
    ]
    
    if filtered_df.empty:
        return None, "No data available for selected period"
    
    # Create dashboard
    st.subheader(f"📊 Monthly SMS Dashboard - {month}/{year}")
    
    # KPI Tiles
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Hazards", len(filtered_df))
    
    with col2:
        high_risk = len(filtered_df[filtered_df['Initial Risk Level'] == 'High'])
        st.metric("High-Risk Hazards", high_risk)
    
    with col3:
        caps_pending = len(filtered_df[filtered_df['CAP Required'] == 'Yes'])
        st.metric("CAPs Required", caps_pending)
    
    with col4:
        wet_lease = len(filtered_df[filtered_df['Wet Lease'] == 'Yes'])
        wet_lease_pct = (wet_lease / len(filtered_df) * 100) if len(filtered_df) > 0 else 0
        st.metric("Wet Lease Incidents", f"{wet_lease_pct:.1f}%")
    
    st.divider()
    
    # Charts
    col1, col2 = st.columns(2)
    
    with col1:
        # Hazards by Location
        if 'Location' in filtered_df.columns:
            location_counts = filtered_df['Location'].value_counts()
            fig1 = px.pie(values=location_counts.values, 
                         names=location_counts.index,
                         title="Hazards by Location",
                         hole=0.3)
            st.plotly_chart(fig1, use_container_width=True)
    
    with col2:
        # Risk Level Distribution
        if 'Initial Risk Level' in filtered_df.columns:
            risk_counts = filtered_df['Initial Risk Level'].value_counts()
            fig2 = px.bar(x=risk_counts.index, y=risk_counts.values,
                         title="Risk Level Distribution",
                         color=risk_counts.index,
                         color_discrete_map={'High': 'red', 'Medium': 'orange', 'Low': 'green'})
            st.plotly_chart(fig2, use_container_width=True)
    
    # CAP Effectiveness
    st.subheader("📈 CAP Effectiveness")
    
    if 'Target Date' in filtered_df.columns and 'Status' in filtered_df.columns:
        cap_df = filtered_df[filtered_df['CAP Required'] == 'Yes'].copy()
        if not cap_df.empty:
            cap_df['Days Overdue'] = cap_df.apply(lambda x: 
                (datetime.now() - datetime.strptime(x['Target Date'], '%d-%m-%Y')).days 
                if pd.notnull(x['Target Date']) and x['Target Date'] != '' else 0, axis=1)
            
            cap_df['Status'] = cap_df['Days Overdue'].apply(
                lambda x: 'Overdue' if x > 0 else ('Due Soon' if x >= -7 else 'On Track')
            )
            
            status_counts = cap_df['Status'].value_counts()
            fig3 = px.bar(x=status_counts.index, y=status_counts.values,
                         title="CAP Status Distribution",
                         color=status_counts.index,
                         color_discrete_map={'Overdue': 'red', 'Due Soon': 'orange', 'On Track': 'green'})
            st.plotly_chart(fig3, use_container_width=True)
    
    # AI Insights
    st.subheader("🤖 AI Safety Insights")
    
    insights = generate_ai_insights(filtered_df)
    st.info(insights)
    
    return filtered_df, ""

def generate_ai_insights(df):
    """Generate AI-powered safety insights"""
    
    insights = []
    
    # Calculate trends
    if len(df) > 5:
        monthly_avg = len(df) / 30  # Approximate daily rate
        insights.append(f"📈 **Monthly Hazard Rate**: {len(df)} hazards reported ({monthly_avg:.1f} per day)")
    
    # Check for high-risk patterns
    high_risk_df = df[df['Initial Risk Level'] == 'High']
    if not high_risk_df.empty:
        common_location = high_risk_df['Location'].mode()[0] if 'Location' in high_risk_df.columns else ""
        insights.append(f"⚠️ **High-Risk Alert**: {len(high_risk_df)} high-risk hazards identified")
        if common_location:
            insights.append(f"   • Most common location: {common_location}")
    
    # Wet lease analysis
    wet_lease_df = df[df['Wet Lease'] == 'Yes']
    if not wet_lease_df.empty:
        wet_lease_pct = (len(wet_lease_df) / len(df) * 100)
        insights.append(f"✈️ **Wet Lease Operations**: {wet_lease_pct:.1f}% of hazards involve wet lease aircraft")
    
    # CAP analysis
    cap_required = df[df['CAP Required'] == 'Yes']
    if not cap_required.empty:
        insights.append(f"🛠️ **Corrective Actions**: {len(cap_required)} CAPs required ({len(cap_required)/len(df)*100:.0f}% of cases)")
    
    # Location analysis
    if 'Location' in df.columns:
        top_location = df['Location'].mode()[0] if not df['Location'].mode().empty else ""
        if top_location:
            location_count = len(df[df['Location'] == top_location])
            insights.append(f"📍 **Top Hazard Location**: {top_location} ({location_count} incidents)")
    
    if not insights:
        insights = ["📊 No significant patterns detected in the current data."]
    
    return "\n\n".join(insights)

def main():
    st.set_page_config(
        page_title="AirSial SMS Scanner & Dashboard",
        page_icon="✈️",
        layout="wide"
    )
    
    st.title("✈️ AirSial SMS Form Scanner & Dashboard Generator")
    st.markdown("### Automated Hazard Report Processing System")
    
    # Sidebar
    with st.sidebar:
        st.header("📋 Navigation")
        app_mode = st.radio(
            "Select Mode:",
            ["📤 Upload & Extract", "📊 Generate Dashboard", "📁 View Data"]
        )
        
        st.divider()
        st.info("**Instructions:**\n1. Upload forms (images/PDFs)\n2. Extract data\n3. Download Excel\n4. Generate dashboard")
        
        if st.session_state.extracted_data:
            st.success(f"✅ {len(st.session_state.extracted_data)} reports processed")
    
    if app_mode == "📤 Upload & Extract":
        st.header("📤 Upload Hazard Reports")
        
        uploaded_files = st.file_uploader(
            "Upload forms (PNG, JPG, PDF)",
            type=['png', 'jpg', 'jpeg', 'pdf'],
            accept_multiple_files=True
        )
        
        if uploaded_files:
            st.success(f"📁 {len(uploaded_files)} files uploaded")
            
            if st.button("🚀 Extract Data from All Files", type="primary"):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i, uploaded_file in enumerate(uploaded_files):
                    status_text.text(f"Processing file {i+1}/{len(uploaded_files)}...")
                    
                    try:
                        # Process file
                        if uploaded_file.type == "application/pdf":
                            images = pdf2image.convert_from_bytes(uploaded_file.read())
                            image = images[0]
                        else:
                            image = Image.open(uploaded_file)
                        
                        # Preprocess and extract text
                        processed_image = preprocess_image(image)
                        text = extract_text_with_paddleocr(processed_image)
                        
                        # Extract form data
                        form_data = extract_form_fields(text)
                        
                        # Store in session
                        st.session_state.extracted_data.append({
                            'filename': uploaded_file.name,
                            'data': form_data,
                            'text_preview': text[:500] + "..." if len(text) > 500 else text
                        })
                        
                    except Exception as e:
                        st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                    
                    progress_bar.progress((i + 1) / len(uploaded_files))
                
                status_text.text("✅ Processing complete!")
                st.balloons()
        
        # Display extracted data
        if st.session_state.extracted_data:
            st.header("📋 Extracted Data Preview")
            
            for i, item in enumerate(st.session_state.extracted_data):
                with st.expander(f"📄 {item['filename']} - {item['data']['basic_details'].get('report_number', 'N/A')}"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.subheader("Basic Details")
                        for key, value in item['data']['basic_details'].items():
                            st.write(f"**{key.replace('_', ' ').title()}**: {value}")
                    
                    with col2:
                        st.subheader("Risk Assessment")
                        for key, value in item['data']['initial_risk'].items():
                            st.write(f"**{key.replace('_', ' ').title()}**: {value}")
                    
                    st.subheader("Corrective Actions")
                    for key, value in item['data']['cap'].items():
                        st.write(f"**{key.replace('_', ' ').title()}**: {value}")
            
            # Create Excel file
            if st.button("📥 Generate Excel Report", type="primary"):
                wb = create_excel_template()
                
                for item in st.session_state.extracted_data:
                    wb = add_data_to_excel(wb, item['data'])
                
                # Save to temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    wb.save(tmp.name)
                    
                    # Read file for download
                    with open(tmp.name, 'rb') as f:
                        excel_bytes = f.read()
                    
                    os.unlink(tmp.name)
                
                # Store in session for dashboard
                df = pd.read_excel(tmp.name, sheet_name='Raw SMS Data')
                st.session_state.excel_data = df
                
                # Download button
                st.download_button(
                    label="⬇️ Download Excel File",
                    data=excel_bytes,
                    file_name=f"AirSial_SMS_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✅ Excel file generated with all sheets!")
    
    elif app_mode == "📊 Generate Dashboard":
        st.header("📊 Monthly SMS Dashboard Generator")
        
        if st.session_state.excel_data.empty:
            st.warning("⚠️ Please upload and extract data first!")
            return
        
        col1, col2 = st.columns(2)
        
        with col1:
            month = st.selectbox(
                "Select Month",
                range(1, 13),
                format_func=lambda x: datetime(2000, x, 1).strftime('%B')
            )
        
        with col2:
            current_year = datetime.now().year
            year = st.selectbox(
                "Select Year",
                range(current_year - 5, current_year + 1),
                index=5
            )
        
        if st.button("📈 Generate Dashboard", type="primary"):
            with st.spinner("Generating dashboard..."):
                filtered_df, error = create_dashboard(st.session_state.excel_data, month, year)
                
                if error:
                    st.error(error)
                else:
                    # Download options
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Excel download
                        excel_buffer = filtered_df.to_excel(index=False)
                        st.download_button(
                            label="📥 Download Dashboard Data (Excel)",
                            data=excel_buffer,
                            file_name=f"Dashboard_{month}_{year}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    with col2:
                        # PDF download placeholder
                        st.info("PDF export feature requires additional setup")
    
    elif app_mode == "📁 View Data":
        st.header("📁 Extracted Data")
        
        if not st.session_state.extracted_data:
            st.info("No data extracted yet. Please upload files first.")
            return
        
        # Display as dataframe
        rows = []
        for item in st.session_state.extracted_data:
            row = {}
            # Flatten data structure
            for category in ['basic_details', 'initial_risk', 'cap', 'residual_risk', 'administrative']:
                if category in item['data']:
                    for key, value in item['data'][category].items():
                        row[f"{category}_{key}"] = value
            rows.append(row)
        
        if rows:
            df_view = pd.DataFrame(rows)
            st.dataframe(df_view, use_container_width=True)
            
            # Summary statistics
            st.subheader("📊 Summary Statistics")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Reports", len(df_view))
            
            with col2:
                high_risk = len(df_view[df_view['initial_risk_risk_level'] == 'High'])
                st.metric("High Risk Reports", high_risk)
            
            with col3:
                caps = len(df_view[df_view['cap_cap_required'] == 'Yes'])
                st.metric("CAPs Required", caps)

if __name__ == "__main__":
    main()
